# 文件提取器模块
import os
import logging
from typing import Dict, List, Optional, Union
from pathlib import Path
import PyPDF2
from docx import Document
from openpyxl import load_workbook

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class FileExtractor:
    """
    文件提取器类，支持多种文件类型的文本提取
    """
    
    def __init__(self):
        """
        初始化文件提取器
        """
        pass
    
    def extract_text(self, file_path: Union[str, Path]) -> Dict[str, Union[str, List[str]]]:
        """
        提取文件文本内容
        
        Args:
            file_path: 文件路径
            
        Returns:
            包含文件路径和提取文本的字典
        """
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                return {
                    "file_path": str(file_path),
                    "error": "文件不存在"
                }
            
            file_extension = file_path.suffix.lower()
            
            if file_extension == '.pdf':
                return self._extract_pdf(file_path)
            elif file_extension == '.docx':
                return self._extract_docx(file_path)
            elif file_extension == '.xlsx':
                return self._extract_xlsx(file_path)
            elif file_extension == '.txt':
                return self._extract_txt(file_path)
            else:
                return {
                    "file_path": str(file_path),
                    "error": "不支持的文件类型"
                }
        except Exception as e:
            logger.error(f"提取文件 {file_path} 时出错: {str(e)}")
            return {
                "file_path": str(file_path),
                "error": str(e)
            }
    
    def batch_extract(self, directory: Union[str, Path], extensions: Optional[List[str]] = None) -> List[Dict[str, Union[str, List[str]]]]:
        """
        批量提取目录中的文件
        
        Args:
            directory: 目录路径
            extensions: 要提取的文件扩展名列表，默认为 None（提取所有支持的类型）
            
        Returns:
            提取结果列表
        """
        try:
            directory = Path(directory)
            if not directory.exists() or not directory.is_dir():
                return [{"error": "目录不存在或不是目录"}]
            
            if extensions is None:
                extensions = ['.pdf', '.docx', '.xlsx', '.txt']
            
            results = []
            for file_path in directory.iterdir():
                if file_path.is_file() and file_path.suffix.lower() in extensions:
                    result = self.extract_text(file_path)
                    results.append(result)
            
            return results
        except Exception as e:
            logger.error(f"批量提取时出错: {str(e)}")
            return [{"error": str(e)}]
    
    def _extract_pdf(self, file_path: Path) -> Dict[str, str]:
        """
        提取 PDF 文件文本
        
        Args:
            file_path: PDF 文件路径
            
        Returns:
            包含文件路径和提取文本的字典
        """
        try:
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ''
                for page_num in range(len(reader.pages)):
                    page = reader.pages[page_num]
                    text += page.extract_text() + '\n'
                
            return {
                "file_path": str(file_path),
                "text": text.strip()
            }
        except Exception as e:
            logger.error(f"提取 PDF 文件 {file_path} 时出错: {str(e)}")
            return {
                "file_path": str(file_path),
                "error": str(e)
            }
    
    def _extract_docx(self, file_path: Path) -> Dict[str, str]:
        """
        提取 Word 文件文本
        
        Args:
            file_path: Word 文件路径
            
        Returns:
            包含文件路径和提取文本的字典
        """
        try:
            doc = Document(file_path)
            text = ''
            for para in doc.paragraphs:
                text += para.text + '\n'
            
            return {
                "file_path": str(file_path),
                "text": text.strip()
            }
        except Exception as e:
            logger.error(f"提取 Word 文件 {file_path} 时出错: {str(e)}")
            return {
                "file_path": str(file_path),
                "error": str(e)
            }
    
    def _extract_xlsx(self, file_path: Path) -> Dict[str, List[str]]:
        """
        提取 Excel 文件文本
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            包含文件路径和提取文本的字典
        """
        try:
            wb = load_workbook(file_path)
            sheets_data = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                    sheet_text += row_text + '\n'
                
                sheets_data.append(sheet_text)
            
            return {
                "file_path": str(file_path),
                "text": sheets_data
            }
        except Exception as e:
            logger.error(f"提取 Excel 文件 {file_path} 时出错: {str(e)}")
            return {
                "file_path": str(file_path),
                "error": str(e)
            }
    
    def _extract_txt(self, file_path: Path) -> Dict[str, str]:
        """
        提取 TXT 文件文本
        
        Args:
            file_path: TXT 文件路径
            
        Returns:
            包含文件路径和提取文本的字典
        """
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
            
            return {
                "file_path": str(file_path),
                "text": text.strip()
            }
        except Exception as e:
            logger.error(f"提取 TXT 文件 {file_path} 时出错: {str(e)}")
            return {
                "file_path": str(file_path),
                "error": str(e)
            }
